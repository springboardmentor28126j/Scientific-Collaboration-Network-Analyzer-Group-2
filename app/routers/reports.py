from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
import matplotlib.pyplot as plt
from fastapi.responses import FileResponse
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from openpyxl.styles import Font, Alignment, Border, Side
import numpy as np
from openpyxl.utils import get_column_letter
from reportlab.platypus import Image
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook

from app.oauth2 import get_current_user
from app.database import get_db
from app import models

import os
from datetime import datetime
router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)



# =====================================================
# 1. PUBLICATION REPORT
# =====================================================


@router.get("/publication-report")
def publication_report(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):


    base_query = db.query(
        models.Publication
    )


    # ===============================
    # Institution Admin
    # ===============================

    if current_user.role == "institution_admin":


        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id
            ==
            current_user.id
        ).first()


        if institution:

            base_query = (
                base_query
                .join(models.Researcher)
                .filter(
                    models.Researcher.institution
                    ==
                    institution.name
                )
            )



    # ===============================
    # Researcher
    # ===============================

    elif current_user.role == "researcher":


        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id
            ==
            current_user.id
        ).first()


        if researcher:

            base_query = base_query.filter(
                models.Publication.researcher_id
                ==
                researcher.id
            )



    # ===============================
    # System Admin
    # ===============================

    elif current_user.role == "system_admin":

        pass



    data = (
        base_query
        .with_entities(
            models.Publication.publication_year,
            func.count(models.Publication.id)
            .label("total")
        )
        .group_by(
            models.Publication.publication_year
        )
        .order_by(
            models.Publication.publication_year
        )
        .all()
    )



    report=[]



    for year,total in data:


        report.append({

            "year":year,

            "total_publications":total,


            "published":
            base_query.filter(
                models.Publication.publication_year == year,
                models.Publication.status == "Published"
            ).count(),


            "submitted":
            base_query.filter(
                models.Publication.publication_year == year,
                models.Publication.status == "Submitted"
            ).count(),


            "draft":
            base_query.filter(
                models.Publication.publication_year == year,
                models.Publication.status == "Draft"
            ).count(),


            "archived":
            base_query.filter(
                models.Publication.publication_year == year,
                models.Publication.status == "Archived"
            ).count()

        })



    return {

        "report":"Publication Report",

        "table":report,


        "chart":{

            "labels":[
                x["year"]
                for x in report
            ],


            "published":[
                x["published"]
                for x in report
            ],


            "submitted":[
                x["submitted"]
                for x in report
            ],


            "draft":[
                x["draft"]
                for x in report
            ],


            "archived":[
                x["archived"]
                for x in report
            ]

        }

    }


# =====================================================
# PUBLICATION REPORT PDF EXPORT
# =====================================================

@router.get("/publication-report/pdf")
def export_publication_report_pdf(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    base_query = db.query(models.Publication)


    # ROLE FILTER

    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            base_query = (
                base_query
                .join(models.Researcher)
                .filter(
                    models.Researcher.institution
                    ==
                    institution.name
                )
            )


    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id == current_user.id
        ).first()


        if researcher:

            base_query = base_query.filter(
                models.Publication.researcher_id
                ==
                researcher.id
            )


    data = (
        base_query
        .with_entities(
            models.Publication.publication_year,
            func.count(models.Publication.id)
        )
        .group_by(
            models.Publication.publication_year
        )
        .order_by(
            models.Publication.publication_year
        )
        .all()
    )


    report=[]


    for year,total in data:

        report.append({

            "year":year,

            "total_publications":total,

            "published":
            base_query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Published"
            ).count(),

            "submitted":
            base_query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Submitted"
            ).count(),

            "draft":
            base_query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Draft"
            ).count(),

            "archived":
            base_query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Archived"
            ).count()

        })



    # ============================
    # CREATE WEBPAGE STYLE GRAPH
    # ============================


    years=[x["year"] for x in report]

    published=[
        x["published"]
        for x in report
    ]

    submitted=[
        x["submitted"]
        for x in report
    ]

    draft=[
        x["draft"]
        for x in report
    ]

    archived=[
        x["archived"]
        for x in report
    ]


    graph_path="publication_chart.png"


    x=np.arange(len(years))

    width=0.2


    plt.figure(figsize=(8,5))


    plt.bar(
        x-width*1.5,
        published,
        width,
        label="Published"
    )


    plt.bar(
        x-width/2,
        submitted,
        width,
        label="Submitted"
    )


    plt.bar(
        x+width/2,
        draft,
        width,
        label="Draft"
    )


    plt.bar(
        x+width*1.5,
        archived,
        width,
        label="Archived"
    )


    plt.xticks(
        x,
        years
    )


    plt.xlabel(
        "Publication Year"
    )


    plt.ylabel(
        "Publications"
    )


    plt.title(
        "Publication Status Analytics"
    )


    plt.legend()


    plt.grid(
        axis="y",
        alpha=0.3
    )


    plt.tight_layout()


    plt.savefig(
        graph_path,
        dpi=300,
        bbox_inches="tight"
    )


    plt.close()



    # ============================
    # CREATE PDF
    # ============================


    file_path="publication_report.pdf"


    doc=SimpleDocTemplate(
        file_path,
        pagesize=letter
    )


    styles=getSampleStyleSheet()


    title_style=ParagraphStyle(
        "title",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        textColor=colors.HexColor("#1E3C72")
    )



    content=[]


    content.append(
        Paragraph(
            "Scientific Collaboration Network Analyzer",
            title_style
        )
    )


    content.append(
        Spacer(1,15)
    )


    content.append(
        Paragraph(
            "Publication Report",
            styles["Heading2"]
        )
    )


    content.append(
        Paragraph(
            f"Generated Date : {datetime.now().strftime('%d-%m-%Y')}",
            styles["Normal"]
        )
    )


    content.append(
        Paragraph(
            f"User Role : {current_user.role}",
            styles["Normal"]
        )
    )


    content.append(
        Spacer(1,20)
    )



    # SUMMARY

    summary_data=[

        ["Metric","Count"],

        [
            "Total Publications",
            sum(x["total_publications"] for x in report)
        ],

        [
            "Published",
            sum(x["published"] for x in report)
        ],

        [
            "Submitted",
            sum(x["submitted"] for x in report)
        ],

        [
            "Draft",
            sum(x["draft"] for x in report)
        ],

        [
            "Archived",
            sum(x["archived"] for x in report)
        ]

    ]


    summary_table=Table(
        summary_data
    )


    summary_table.setStyle(
        TableStyle([
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3C72")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(0,0),(-1,-1),"CENTER")
        ])
    )


    content.append(
        Paragraph(
            "Summary",
            styles["Heading3"]
        )
    )


    content.append(
        summary_table
    )


    content.append(
        Spacer(1,20)
    )



    # GRAPH


    content.append(
        Paragraph(
            "Publication Analytics",
            styles["Heading3"]
        )
    )


    content.append(
        Image(
            graph_path,
            width=400,
            height=250
        )
    )


    content.append(
        Spacer(1,20)
    )



    # DETAILS TABLE


    table_data=[

        [
            "Year",
            "Total Publications",
            "Published",
            "Submitted",
            "Draft",
            "Archived"
        ]

    ]


    for item in report:

        table_data.append(

            [

                item["year"],
                item["total_publications"],
                item["published"],
                item["submitted"],
                item["draft"],
                item["archived"]

            ]

        )



    table=Table(
        table_data
    )


    table.setStyle(
        TableStyle([
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3C72")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(0,0),(-1,-1),"CENTER")
        ])
    )


    content.append(
        Paragraph(
            "Publication Details",
            styles["Heading3"]
        )
    )


    content.append(
        table
    )



    doc.build(
        content
    )


    return FileResponse(
        file_path,
        filename="publication_report.pdf"
    )

# =====================================================
# PUBLICATION REPORT EXCEL EXPORT
# =====================================================

@router.get("/publication-report/excel")
def export_publication_report_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):


    query = db.query(models.Publication)


    # ============================
    # ROLE FILTER
    # ============================

    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            query = (
                query
                .join(models.Researcher)
                .filter(
                    models.Researcher.institution
                    ==
                    institution.name
                )
            )


    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id == current_user.id
        ).first()


        if researcher:

            query = query.filter(
                models.Publication.researcher_id
                ==
                researcher.id
            )



    publications = query.all()



    report=[]


    years=sorted(
        list(
            set(
                [
                    p.publication_year
                    for p in publications
                ]
            )
        )
    )



    for year in years:

        report.append({

            "year":year,

            "total_publications":
            query.filter(
                models.Publication.publication_year==year
            ).count(),


            "published":
            query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Published"
            ).count(),


            "submitted":
            query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Submitted"
            ).count(),


            "draft":
            query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Draft"
            ).count(),


            "archived":
            query.filter(
                models.Publication.publication_year==year,
                models.Publication.status=="Archived"
            ).count()

        })



    # ============================
    # WORKBOOK
    # ============================


    wb=Workbook()


    ws=wb.active


    ws.title="Publication Report"



    center=Alignment(
        horizontal="center",
        vertical="center"
    )


    thin=Side(
        style="thin"
    )


    border=Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )



    # ============================
    # MAIN HEADER
    # ============================


    ws.merge_cells(
        "A1:F1"
    )


    ws["A1"] = (
        "Scientific Collaboration Network Analyzer"
    )


    ws["A1"].font=Font(
        bold=True,
        size=16
    )


    ws["A1"].alignment=center



    ws.merge_cells(
        "A2:F2"
    )


    ws["A2"]="Publication Report"


    ws["A2"].font=Font(
        bold=True,
        size=14
    )


    ws["A2"].alignment=center



    ws.merge_cells(
        "A3:F3"
    )


    ws["A3"]=(
        f"Generated Date : "
        f"{datetime.now().strftime('%d-%m-%Y')}"
    )


    ws["A3"].alignment=center



    # ============================
    # SUMMARY HEADING
    # ============================


    row=5


    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=2
    )


    ws.cell(
        row=row,
        column=1,
        value="Summary"
    )


    ws.cell(
        row=row,
        column=1
    ).font=Font(
        bold=True,
        size=12
    )


    ws.cell(
        row=row,
        column=1
    ).alignment=center



    row+=1



    summary=[

        [
            "Metric",
            "Count"
        ],

        [
            "Total Publications",
            len(publications)
        ],

        [
            "Published",
            sum(
                x["published"]
                for x in report
            )
        ],

        [
            "Submitted",
            sum(
                x["submitted"]
                for x in report
            )
        ],

        [
            "Draft",
            sum(
                x["draft"]
                for x in report
            )
        ],

        [
            "Archived",
            sum(
                x["archived"]
                for x in report
            )
        ]

    ]



    for item in summary:


        for col,value in enumerate(item,1):


            cell=ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border=border


            if row==6:

                cell.font=Font(
                    bold=True
                )

                cell.alignment=center


        row+=1




    # ============================
    # PUBLICATION DETAILS HEADING
    # ============================


    row+=2



    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=6
    )


    ws.cell(
        row=row,
        column=1,
        value="Publication Details"
    )


    ws.cell(
        row=row,
        column=1
    ).font=Font(
        bold=True,
        size=12
    )


    ws.cell(
        row=row,
        column=1
    ).alignment=center



    row+=1



    headers=[

        "Year",
        "Total Publications",
        "Published",
        "Submitted",
        "Draft",
        "Archived"

    ]



    for col,header in enumerate(headers,1):


        cell=ws.cell(
            row=row,
            column=col,
            value=header
        )


        cell.font=Font(
            bold=True
        )


        cell.alignment=center


        cell.border=border



    row+=1



    for item in report:


        values=[

            item["year"],
            item["total_publications"],
            item["published"],
            item["submitted"],
            item["draft"],
            item["archived"]

        ]



        for col,value in enumerate(values,1):


            cell=ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border=border



        row+=1



    # ============================
    # WIDTH
    # ============================


    widths={

        "A":15,
        "B":22,
        "C":18,
        "D":18,
        "E":15,
        "F":15

    }



    for col,width in widths.items():

        ws.column_dimensions[col].width=width



    # ============================
    # SAVE
    # ============================


    file_path="publication_report.xlsx"


    wb.save(
        file_path
    )



    return FileResponse(
        file_path,
        filename="publication_report.xlsx"
    )
# =====================================================
# 2. RESEARCH REPORT
# =====================================================


@router.get("/research-report")
def research_report(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):


    query = db.query(
        models.Researcher
    )



    if current_user.role == "institution_admin":


        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id
            ==
            current_user.id
        ).first()



        if institution:

            query=query.filter(
                models.Researcher.institution
                ==
                institution.name
            )




    elif current_user.role == "researcher":


        query=query.filter(
            models.Researcher.user_id
            ==
            current_user.id
        )




    elif current_user.role == "system_admin":

        pass




    researchers=query.all()



    report=[]



    for researcher in researchers:



        publications=db.query(
            models.Publication
        ).filter(
            models.Publication.researcher_id
            ==
            researcher.id
        ).count()



        projects=db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.researcher_id
            ==
            researcher.id
        ).count()



        collaborations=db.query(
            models.InstitutionCollaboration
        ).join(
            models.Project
        ).join(
            models.ProjectMember
        ).filter(
            models.ProjectMember.researcher_id
            ==
            researcher.id
        ).count()



        report.append({

            "researcher":
            researcher.full_name,


            "publications":
            publications,


            "projects":
            projects,


            "collaborations":
            collaborations

        })




    return {


        "report":"Research Report",


        "table":report,


        "chart":{

            "labels":[
                x["researcher"]
                for x in report
            ],


            "values":[
                x["publications"]
                for x in report
            ]

        }

    }

# ==========================================
# Research Report PDF Export
# ==========================================

@router.get("/research-report/pdf")
def export_research_report_pdf(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # ============================
    # GET RESEARCHERS BASED ON ROLE
    # ============================

    query = db.query(models.Researcher)


    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            query = query.filter(
                models.Researcher.institution
                ==
                institution.name
            )


    elif current_user.role == "researcher":

        query = query.filter(
            models.Researcher.user_id
            ==
            current_user.id
        )


    # system_admin gets all data


    researchers = query.all()



    # ============================
    # PREPARE REPORT DATA
    # ============================

    report = []


    for researcher in researchers:


        publications = db.query(
            models.Publication
        ).filter(
            models.Publication.researcher_id
            ==
            researcher.id
        ).count()



        projects = db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.researcher_id
            ==
            researcher.id
        ).count()



        collaborations = (
            db.query(
                models.InstitutionCollaboration
            )
            .join(models.Project)
            .join(models.ProjectMember)
            .filter(
                models.ProjectMember.researcher_id
                ==
                researcher.id
            )
            .count()
        )


        report.append({

            "researcher": researcher.full_name,

            "publications": publications,

            "projects": projects,

            "collaborations": collaborations

        })



    # ============================
    # CREATE GRAPH
    # ============================


    names = [
        x["researcher"]
        for x in report
    ]


    publications = [
        x["publications"]
        for x in report
    ]


    projects = [
        x["projects"]
        for x in report
    ]


    collaborations = [
        x["collaborations"]
        for x in report
    ]



    graph_path = "research_chart.png"


    y = np.arange(
        len(names)
    )


    height = 0.25



    plt.figure(
        figsize=(8,5)
    )


    plt.barh(
        y-height,
        publications,
        height,
        label="Publications"
    )


    plt.barh(
        y,
        projects,
        height,
        label="Projects"
    )


    plt.barh(
        y+height,
        collaborations,
        height,
        label="Collaborations"
    )



    plt.yticks(
        y,
        names
    )


    plt.xlabel(
        "Count"
    )


    plt.ylabel(
        "Researchers"
    )


    plt.title(
        "Research Publications, Projects and Collaborations"
    )


    plt.legend()



    plt.grid(
        axis="x",
        alpha=0.3
    )


    plt.tight_layout()


    plt.savefig(
        graph_path,
        dpi=300,
        bbox_inches="tight"
    )


    plt.close()



    # ============================
    # CREATE PDF
    # ============================


    file_path = "research_report.pdf"



    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter
    )


    styles = getSampleStyleSheet()



    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        textColor=colors.HexColor("#1E3C72")
    )



    content = []



    # Header

    content.append(
        Paragraph(
            "Scientific Collaboration Network Analyzer",
            title_style
        )
    )


    content.append(
        Spacer(1,15)
    )


    content.append(
        Paragraph(
            "Research Report",
            styles["Heading2"]
        )
    )


    content.append(
        Paragraph(
            f"Generated Date : {datetime.now().strftime('%d-%m-%Y')}",
            styles["Normal"]
        )
    )


    content.append(
        Paragraph(
            f"User Role : {current_user.role}",
            styles["Normal"]
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # SUMMARY
    # ============================


    summary_data = [

        ["Metric","Count"],

        [
            "Total Researchers",
            len(report)
        ],

        [
            "Total Publications",
            sum(
                x["publications"]
                for x in report
            )
        ],

        [
            "Total Projects",
            sum(
                x["projects"]
                for x in report
            )
        ],

        [
            "Total Collaborations",
            sum(
                x["collaborations"]
                for x in report
            )
        ]

    ]



    summary_table = Table(
        summary_data
    )


    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])
    )


    content.append(
        Paragraph(
            "Summary",
            styles["Heading3"]
        )
    )


    content.append(
        summary_table
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # GRAPH
    # ============================


    content.append(
        Paragraph(
            "Researcher Performance Analytics",
            styles["Heading3"]
        )
    )


    content.append(
        Image(
            graph_path,
            width=400,
            height=250
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # RESEARCH TABLE
    # ============================


    table_data = [

        [
            "Rank",
            "Researcher Name",
            "Publications",
            "Projects",
            "Collaborations"
        ]

    ]



    rank = 1


    for item in report:

        table_data.append(

            [
                rank,
                item["researcher"],
                item["publications"],
                item["projects"],
                item["collaborations"]
            ]

        )

        rank += 1



    research_table = Table(
        table_data
    )



    research_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])

    )



    content.append(
        Paragraph(
            "Researcher Performance Table",
            styles["Heading3"]
        )
    )


    content.append(
        research_table
    )


    content.append(
        Spacer(1,20)
    )



    # Footer

    content.append(
        Paragraph(
            "Generated by Scientific Collaboration Network Analyzer",
            styles["Normal"]
        )
    )



    doc.build(
        content
    )



    return FileResponse(
        file_path,
        filename="research_report.pdf"
    )

# ==========================================
# Research Report Excel Export
# ==========================================

@router.get("/research-report/excel")
def export_research_report_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(models.Researcher)


    # ============================
    # ROLE FILTER
    # ============================

    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            query = query.filter(
                models.Researcher.institution
                ==
                institution.name
            )


    elif current_user.role == "researcher":

        query = query.filter(
            models.Researcher.user_id
            ==
            current_user.id
        )


    researchers = query.all()



    # ============================
    # CREATE WORKBOOK
    # ============================

    wb = Workbook()

    ws = wb.active

    ws.title = "Research Report"



    # ============================
    # STYLES
    # ============================

    title_font = Font(
        bold=True,
        size=16
    )


    heading_font = Font(
        bold=True,
        size=12
    )


    header_font = Font(
        bold=True
    )


    center = Alignment(
        horizontal="center",
        vertical="center"
    )


    thin = Side(
        style="thin"
    )


    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )



    # ============================
    # MAIN HEADER
    # ============================

    ws.merge_cells(
        "A1:E1"
    )

    ws["A1"] = (
        "Scientific Collaboration "
        "Network Analyzer"
    )


    ws["A1"].font = title_font

    ws["A1"].alignment = center



    ws.merge_cells(
        "A2:E2"
    )


    ws["A2"] = "Research Report"


    ws["A2"].font = Font(
        bold=True,
        size=14
    )


    ws["A2"].alignment = center



    ws.merge_cells(
        "A3:E3"
    )


    ws["A3"] = (
        f"Generated Date : "
        f"{datetime.now().strftime('%d-%m-%Y')}"
    )


    ws["A3"].alignment = center



    # ============================
    # SUMMARY SECTION
    # ============================

    row = 5


    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=2
    )


    ws.cell(
        row=row,
        column=1,
        value="Summary"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    summary = [

        ["Metric","Count"],

        [
            "Researchers",
            len(researchers)
        ],

        [
            "Publications",
            db.query(models.Publication).count()
        ],

        [
            "Projects",
            db.query(models.Project).count()
        ],

        [
            "Collaborations",
            db.query(
                models.InstitutionCollaboration
            ).count()
        ]

    ]



    for item in summary:

        for col,value in enumerate(item,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )

            cell.border = border


            if row == 6:

                cell.font = header_font

                cell.alignment = center


        row += 1




    # ============================
    # RESEARCHER TABLE
    # ============================


    row += 2


    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=5
    )


    ws.cell(
        row=row,
        column=1,
        value="Researcher Performance Table"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    headers = [

        "Rank",
        "Researcher Name",
        "Publications",
        "Projects",
        "Collaborations"

    ]



    for col,header in enumerate(headers,1):

        cell = ws.cell(
            row=row,
            column=col,
            value=header
        )


        cell.font = header_font

        cell.alignment = center

        cell.border = border



    row += 1


    rank = 1



    for researcher in researchers:


        publications = db.query(
            models.Publication
        ).filter(
            models.Publication.researcher_id
            ==
            researcher.id
        ).count()



        projects = db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.researcher_id
            ==
            researcher.id
        ).count()



        collaborations = (
            db.query(
                models.InstitutionCollaboration
            )
            .join(models.Project)
            .join(models.ProjectMember)
            .filter(
                models.ProjectMember.researcher_id
                ==
                researcher.id
            )
            .count()
        )



        data = [

            rank,
            researcher.full_name,
            publications,
            projects,
            collaborations

        ]



        for col,value in enumerate(data,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )

            cell.border = border


        rank += 1

        row += 1



    # ============================
    # COLUMN WIDTH
    # ============================

    widths = {

        "A":15,
        "B":30,
        "C":18,
        "D":15,
        "E":20

    }


    for col,width in widths.items():

        ws.column_dimensions[col].width = width



    # ============================
    # SAVE
    # ============================


    file_path = "research_report.xlsx"


    wb.save(
        file_path
    )


    return FileResponse(
        file_path,
        filename="research_report.xlsx"
    )
# =====================================================
# 3. COLLABORATION REPORT
# =====================================================


@router.get("/collaboration-report")
def collaboration_report(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):


    query=db.query(
        models.Project
    )



    if current_user.role=="institution_admin":


        institution=db.query(
            models.Institution
        ).filter(
            models.Institution.user_id
            ==
            current_user.id
        ).first()



        if institution:

            query=query.filter(
                models.Project.institution_id
                ==
                institution.id
            )




    elif current_user.role=="researcher":


        researcher=db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id
            ==
            current_user.id
        ).first()



        if researcher:


            query=(
                query
                .join(models.ProjectMember)
                .filter(
                    models.ProjectMember.researcher_id
                    ==
                    researcher.id
                )
            )




    elif current_user.role=="system_admin":

        pass




    projects=query.all()


    report=[]



    for project in projects:


        members=db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.project_id
            ==
            project.id
        ).count()



        report.append({

            "project":
            project.project_name,


            "institution":
            project.institution.name
            if project.institution
            else "N/A",


            "researchers":
            members

        })



    return {


        "report":"Collaboration Report",


        "table":report,


        "chart":{

            "labels":[
                x["project"]
                for x in report
            ],


            "values":[
                x["researchers"]
                for x in report
            ]

        }

    }


# ==========================================
# Collaboration Report PDF Export
# ==========================================

@router.get("/collaboration-report/pdf")
def export_collaboration_report_pdf(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # ============================
    # GET PROJECT DATA
    # ============================

    query = db.query(models.Project)


    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            query = query.filter(
                models.Project.institution_id
                ==
                institution.id
            )


    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id == current_user.id
        ).first()


        if researcher:

            query = (
                query
                .join(models.ProjectMember)
                .filter(
                    models.ProjectMember.researcher_id
                    ==
                    researcher.id
                )
            )


    projects = query.all()



    # ============================
    # PREPARE REPORT DATA
    # ============================

    report = []


    for project in projects:


        researchers = db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.project_id
            ==
            project.id
        ).count()



        report.append({

            "project":
                project.project_name,


            "institution":
                project.institution.name
                if project.institution
                else "N/A",


            "researchers":
                researchers

        })



    # ============================
    # CREATE GRAPH
    # ============================

    names = [
        x["project"]
        for x in report
    ]


    values = [
        x["researchers"]
        for x in report
    ]



    graph_path = "collaboration_chart.png"



    plt.figure(
        figsize=(8,5)
    )


    plt.bar(
        names,
        values
    )


    plt.xlabel(
        "Projects"
    )


    plt.ylabel(
        "Researchers"
    )


    plt.title(
        "Project Collaboration Analysis"
    )


    plt.xticks(
        rotation=45,
        ha="right"
    )


    plt.tight_layout()


    plt.savefig(
        graph_path,
        dpi=300,
        bbox_inches="tight"
    )


    plt.close()



    # ============================
    # CREATE PDF
    # ============================


    file_path = "collaboration_report.pdf"



    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter
    )



    styles = getSampleStyleSheet()



    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        textColor=colors.HexColor("#1E3C72")
    )



    content = []



    # HEADER

    content.append(
        Paragraph(
            "Scientific Collaboration Network Analyzer",
            title_style
        )
    )


    content.append(
        Spacer(1,15)
    )


    content.append(
        Paragraph(
            "Collaboration Report",
            styles["Heading2"]
        )
    )


    content.append(
        Paragraph(
            f"Generated Date : {datetime.now().strftime('%d-%m-%Y')}",
            styles["Normal"]
        )
    )


    content.append(
        Paragraph(
            f"User Role : {current_user.role}",
            styles["Normal"]
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # SUMMARY TABLE
    # ============================


    summary_data = [

        ["Metric","Count"],

        [
            "Total Projects",
            len(report)
        ],

        [
            "Total Researchers Involved",
            sum(
                x["researchers"]
                for x in report
            )
        ]

    ]



    summary_table = Table(
        summary_data
    )


    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])
    )



    content.append(
        Paragraph(
            "Summary",
            styles["Heading3"]
        )
    )


    content.append(
        summary_table
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # GRAPH
    # ============================


    content.append(
        Paragraph(
            "Collaboration Analytics",
            styles["Heading3"]
        )
    )


    content.append(
        Image(
            graph_path,
            width=400,
            height=250
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # DETAILS TABLE
    # ============================


    table_data = [

        [
            "Rank",
            "Project",
            "Institution",
            "Researchers"
        ]

    ]


    rank = 1


    for item in report:

        table_data.append(

            [

                rank,

                item["project"],

                item["institution"],

                item["researchers"]

            ]

        )

        rank += 1



    collaboration_table = Table(
        table_data
    )



    collaboration_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])

    )



    content.append(
        Paragraph(
            "Collaboration Details",
            styles["Heading3"]
        )
    )


    content.append(
        collaboration_table
    )


    content.append(
        Spacer(1,20)
    )



    content.append(
        Paragraph(
            "Generated by Scientific Collaboration Network Analyzer",
            styles["Normal"]
        )
    )



    doc.build(
        content
    )



    return FileResponse(
        file_path,
        filename="collaboration_report.pdf"
    )
# ==========================================
# Collaboration Report Excel Export
# ==========================================

@router.get("/collaboration-report/excel")
def export_collaboration_report_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # ============================
    # GET PROJECT DATA
    # ============================

    query = db.query(models.Project)



    if current_user.role == "institution_admin":

        institution = db.query(
            models.Institution
        ).filter(
            models.Institution.user_id == current_user.id
        ).first()


        if institution:

            query = query.filter(
                models.Project.institution_id
                ==
                institution.id
            )



    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id == current_user.id
        ).first()


        if researcher:

            query = (
                query
                .join(models.ProjectMember)
                .filter(
                    models.ProjectMember.researcher_id
                    ==
                    researcher.id
                )
            )



    projects = query.all()



    # ============================
    # PREPARE WORKBOOK
    # ============================

    wb = Workbook()

    ws = wb.active

    ws.title = "Collaboration Report"



    # ============================
    # STYLES
    # ============================

    title_font = Font(
        bold=True,
        size=16
    )


    heading_font = Font(
        bold=True,
        size=12
    )


    header_font = Font(
        bold=True
    )


    center = Alignment(
        horizontal="center",
        vertical="center"
    )


    thin = Side(
        style="thin"
    )


    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )



    # ============================
    # MAIN HEADER
    # ============================

    ws.merge_cells(
        "A1:D1"
    )


    ws["A1"] = (
        "Scientific Collaboration "
        "Network Analyzer"
    )


    ws["A1"].font = title_font

    ws["A1"].alignment = center



    ws.merge_cells(
        "A2:D2"
    )


    ws["A2"] = "Collaboration Report"


    ws["A2"].font = Font(
        bold=True,
        size=14
    )


    ws["A2"].alignment = center



    ws.merge_cells(
        "A3:D3"
    )


    ws["A3"] = (
        f"Generated Date : "
        f"{datetime.now().strftime('%d-%m-%Y')}"
    )


    ws["A3"].alignment = center



    # ============================
    # SUMMARY
    # ============================

    row = 5


    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=2
    )


    ws.cell(
        row=row,
        column=1,
        value="Summary"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    summary = [

        [
            "Metric",
            "Count"
        ],

        [
            "Total Projects",
            len(projects)
        ],

        [
            "Total Researchers Involved",
            db.query(
                models.ProjectMember
            ).count()
        ]

    ]



    for data in summary:

        for col,value in enumerate(data,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border = border


            if row == 6:

                cell.font = header_font

                cell.alignment = center


        row += 1



    # ============================
    # COLLABORATION DETAILS
    # ============================

    row += 2



    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=4
    )


    ws.cell(
        row=row,
        column=1,
        value="Collaboration Details"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    headers = [

        "Rank",
        "Project",
        "Institution",
        "Researchers"

    ]



    for col,header in enumerate(headers,1):

        cell = ws.cell(
            row=row,
            column=col,
            value=header
        )


        cell.font = header_font

        cell.alignment = center

        cell.border = border



    row += 1


    rank = 1



    for project in projects:


        researchers = db.query(
            models.ProjectMember
        ).filter(
            models.ProjectMember.project_id
            ==
            project.id
        ).count()



        data = [

            rank,

            project.project_name,

            project.institution.name
            if project.institution
            else "N/A",

            researchers

        ]



        for col,value in enumerate(data,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border = border



        rank += 1

        row += 1



    # ============================
    # COLUMN WIDTH
    # ============================

    widths = {

        "A":15,
        "B":30,
        "C":30,
        "D":18

    }



    for col,width in widths.items():

        ws.column_dimensions[col].width = width



    # ============================
    # SAVE FILE
    # ============================

    file_path = "collaboration_report.xlsx"


    wb.save(
        file_path
    )


    return FileResponse(
        file_path,
        filename="collaboration_report.xlsx"
    )


# =====================================================
# 4. INSTITUTION REPORT
# =====================================================


@router.get("/institution-report")
def institution_report(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):


    query=db.query(
        models.Institution
    )



    if current_user.role=="institution_admin":


        query=query.filter(
            models.Institution.user_id
            ==
            current_user.id
        )




    elif current_user.role=="researcher":


        researcher=db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id
            ==
            current_user.id
        ).first()



        if researcher:


            query=query.filter(
                models.Institution.name
                ==
                researcher.institution
            )




    elif current_user.role=="system_admin":

        pass




    institutions=query.all()


    report=[]



    for institution in institutions:


        researchers=db.query(
            models.Researcher
        ).filter(
            models.Researcher.institution
            ==
            institution.name
        ).count()



        publications=db.query(
            models.Publication
        ).join(
            models.Researcher
        ).filter(
            models.Researcher.institution
            ==
            institution.name
        ).count()



        projects=db.query(
            models.Project
        ).filter(
            models.Project.institution_id
            ==
            institution.id
        ).count()



        report.append({

            "institution":
            institution.name,


            "researchers":
            researchers,


            "publications":
            publications,


            "projects":
            projects

        })



    return {


        "report":"Institution Report",


        "table":report,


        "chart":{

            "labels":[
                x["institution"]
                for x in report
            ],


            "values":[
                x["publications"]
                for x in report
            ]

        }

    }

# ==========================================
# Institution Report PDF Export
# ==========================================

@router.get("/institution-report/pdf")
def export_institution_report_pdf(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # ============================
    # GET INSTITUTION DATA
    # ============================

    query = db.query(models.Institution)



    if current_user.role == "institution_admin":

        query = query.filter(
            models.Institution.user_id
            ==
            current_user.id
        )



    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id
            ==
            current_user.id
        ).first()


        if researcher:

            query = query.filter(
                models.Institution.name
                ==
                researcher.institution
            )



    institutions = query.all()



    # ============================
    # PREPARE REPORT DATA
    # ============================

    report = []



    for institution in institutions:


        researchers = db.query(
            models.Researcher
        ).filter(
            models.Researcher.institution
            ==
            institution.name
        ).count()



        publications = (
            db.query(
                models.Publication
            )
            .join(models.Researcher)
            .filter(
                models.Researcher.institution
                ==
                institution.name
            )
            .count()
        )



        projects = db.query(
            models.Project
        ).filter(
            models.Project.institution_id
            ==
            institution.id
        ).count()



        report.append({

            "institution":
                institution.name,

            "researchers":
                researchers,

            "publications":
                publications,

            "projects":
                projects

        })



    # ============================
    # CREATE GRAPH
    # ============================

    names = [
        x["institution"]
        for x in report
    ]


    publications = [
        x["publications"]
        for x in report
    ]


    projects = [
        x["projects"]
        for x in report
    ]


    researchers = [
        x["researchers"]
        for x in report
    ]



    graph_path = "institution_chart.png"



    x = np.arange(
        len(names)
    )


    width = 0.25



    plt.figure(
        figsize=(8,5)
    )



    plt.bar(
        x-width,
        researchers,
        width,
        label="Researchers"
    )


    plt.bar(
        x,
        publications,
        width,
        label="Publications"
    )


    plt.bar(
        x+width,
        projects,
        width,
        label="Projects"
    )



    plt.xticks(
        x,
        names,
        rotation=45,
        ha="right"
    )


    plt.ylabel(
        "Count"
    )


    plt.title(
        "Institution Research Analytics"
    )


    plt.legend()


    plt.tight_layout()



    plt.savefig(
        graph_path,
        dpi=300,
        bbox_inches="tight"
    )


    plt.close()



    # ============================
    # CREATE PDF
    # ============================

    file_path = "institution_report.pdf"



    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter
    )



    styles = getSampleStyleSheet()



    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        textColor=colors.HexColor("#1E3C72")
    )



    content = []



    content.append(
        Paragraph(
            "Scientific Collaboration Network Analyzer",
            title_style
        )
    )


    content.append(
        Spacer(1,15)
    )


    content.append(
        Paragraph(
            "Institution Report",
            styles["Heading2"]
        )
    )


    content.append(
        Paragraph(
            f"Generated Date : {datetime.now().strftime('%d-%m-%Y')}",
            styles["Normal"]
        )
    )


    content.append(
        Paragraph(
            f"User Role : {current_user.role}",
            styles["Normal"]
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # SUMMARY TABLE
    # ============================


    summary_data = [

        [
            "Metric",
            "Count"
        ],

        [
            "Total Institutions",
            len(report)
        ],

        [
            "Total Researchers",
            sum(
                x["researchers"]
                for x in report
            )
        ],

        [
            "Total Publications",
            sum(
                x["publications"]
                for x in report
            )
        ],

        [
            "Total Projects",
            sum(
                x["projects"]
                for x in report
            )
        ]

    ]



    summary_table = Table(
        summary_data
    )



    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])
    )



    content.append(
        Paragraph(
            "Summary",
            styles["Heading3"]
        )
    )


    content.append(
        summary_table
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # GRAPH
    # ============================


    content.append(
        Paragraph(
            "Institution Analytics",
            styles["Heading3"]
        )
    )


    content.append(
        Image(
            graph_path,
            width=400,
            height=250
        )
    )


    content.append(
        Spacer(1,20)
    )



    # ============================
    # DETAILS TABLE
    # ============================


    table_data = [

        [
            "Rank",
            "Institution",
            "Researchers",
            "Publications",
            "Projects"
        ]

    ]



    rank = 1


    for item in report:

        table_data.append(

            [

                rank,

                item["institution"],

                item["researchers"],

                item["publications"],

                item["projects"]

            ]

        )

        rank += 1



    institution_table = Table(
        table_data
    )



    institution_table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.HexColor("#1E3C72")
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])

    )



    content.append(
        Paragraph(
            "Institution Details",
            styles["Heading3"]
        )
    )


    content.append(
        institution_table
    )


    content.append(
        Spacer(1,20)
    )



    content.append(
        Paragraph(
            "Generated by Scientific Collaboration Network Analyzer",
            styles["Normal"]
        )
    )



    doc.build(
        content
    )



    return FileResponse(
        file_path,
        filename="institution_report.pdf"
    )

# ==========================================
# Institution Report Excel Export
# ==========================================

@router.get("/institution-report/excel")
def export_institution_report_excel(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # ============================
    # GET INSTITUTION DATA
    # ============================

    query = db.query(models.Institution)



    if current_user.role == "institution_admin":

        query = query.filter(
            models.Institution.user_id
            ==
            current_user.id
        )



    elif current_user.role == "researcher":

        researcher = db.query(
            models.Researcher
        ).filter(
            models.Researcher.user_id
            ==
            current_user.id
        ).first()


        if researcher:

            query = query.filter(
                models.Institution.name
                ==
                researcher.institution
            )



    institutions = query.all()



    # ============================
    # CREATE WORKBOOK
    # ============================

    wb = Workbook()

    ws = wb.active

    ws.title = "Institution Report"



    # ============================
    # STYLES
    # ============================

    title_font = Font(
        bold=True,
        size=16
    )


    heading_font = Font(
        bold=True,
        size=12
    )


    header_font = Font(
        bold=True
    )


    center = Alignment(
        horizontal="center",
        vertical="center"
    )


    thin = Side(
        style="thin"
    )


    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )



    # ============================
    # MAIN HEADER
    # ============================

    ws.merge_cells(
        "A1:E1"
    )


    ws["A1"] = (
        "Scientific Collaboration "
        "Network Analyzer"
    )


    ws["A1"].font = title_font

    ws["A1"].alignment = center



    ws.merge_cells(
        "A2:E2"
    )


    ws["A2"] = "Institution Report"


    ws["A2"].font = Font(
        bold=True,
        size=14
    )


    ws["A2"].alignment = center



    ws.merge_cells(
        "A3:E3"
    )


    ws["A3"] = (
        f"Generated Date : "
        f"{datetime.now().strftime('%d-%m-%Y')}"
    )


    ws["A3"].alignment = center



    # ============================
    # PREPARE DATA
    # ============================

    report = []



    for institution in institutions:


        researchers = db.query(
            models.Researcher
        ).filter(
            models.Researcher.institution
            ==
            institution.name
        ).count()



        publications = (
            db.query(
                models.Publication
            )
            .join(models.Researcher)
            .filter(
                models.Researcher.institution
                ==
                institution.name
            )
            .count()
        )



        projects = db.query(
            models.Project
        ).filter(
            models.Project.institution_id
            ==
            institution.id
        ).count()



        report.append({

            "institution":
                institution.name,

            "researchers":
                researchers,

            "publications":
                publications,

            "projects":
                projects

        })



    # ============================
    # SUMMARY
    # ============================

    row = 5



    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=2
    )


    ws.cell(
        row=row,
        column=1,
        value="Summary"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    summary = [

        [
            "Metric",
            "Count"
        ],

        [
            "Institutions",
            len(report)
        ],

        [
            "Researchers",
            sum(
                x["researchers"]
                for x in report
            )
        ],

        [
            "Publications",
            sum(
                x["publications"]
                for x in report
            )
        ],

        [
            "Projects",
            sum(
                x["projects"]
                for x in report
            )
        ]

    ]



    for data in summary:

        for col,value in enumerate(data,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border = border



            if row == 6:

                cell.font = header_font

                cell.alignment = center


        row += 1



    # ============================
    # INSTITUTION DETAILS
    # ============================

    row += 2



    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=5
    )


    ws.cell(
        row=row,
        column=1,
        value="Institution Details"
    )


    ws.cell(
        row=row,
        column=1
    ).font = heading_font


    ws.cell(
        row=row,
        column=1
    ).alignment = center



    row += 1



    headers = [

        "Rank",
        "Institution",
        "Researchers",
        "Publications",
        "Projects"

    ]



    for col,header in enumerate(headers,1):

        cell = ws.cell(
            row=row,
            column=col,
            value=header
        )


        cell.font = header_font

        cell.alignment = center

        cell.border = border



    row += 1



    rank = 1



    for item in report:


        data = [

            rank,

            item["institution"],

            item["researchers"],

            item["publications"],

            item["projects"]

        ]



        for col,value in enumerate(data,1):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )


            cell.border = border



        rank += 1

        row += 1



    # ============================
    # COLUMN WIDTH
    # ============================

    widths = {

        "A":15,

        "B":35,

        "C":18,

        "D":18,

        "E":15

    }



    for col,width in widths.items():

        ws.column_dimensions[col].width = width



    # ============================
    # SAVE
    # ============================

    file_path = "institution_report.xlsx"


    wb.save(
        file_path
    )



    return FileResponse(

        file_path,

        filename="institution_report.xlsx"

    )
