from collections import Counter, defaultdict
from datetime import datetime, timezone

from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app import crud, models, auth
from app.database import get_db
from app.notification_service import notify_all_users
from app.permissions import current_user, is_system_admin, require_roles, require_system_admin

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
    dependencies=[Depends(auth.require_authenticated)]
)


def _allowed_institution_id(user: models.User, db: Session) -> int | None:
    if is_system_admin(user) or user.role.lower() in {"publisher", "reviewer"}:
        return None
    if user.role.lower() == "institution admin":
        return user.institution_id
    researcher = db.query(models.Researcher).filter(models.Researcher.id == user.researcher_id).first() if user.researcher_id else None
    return researcher.institution_id if researcher else None


def _require_report_scope(institution_id: int, user: models.User, db: Session) -> None:
    allowed_id = _allowed_institution_id(user, db)
    if allowed_id is not None and allowed_id != institution_id:
        raise HTTPException(status_code=403, detail="This report is outside your assigned workspace")

@router.get("/publications-count")
def publications_count(db: Session = Depends(get_db)):
    return {"publications_count": crud.count_publications(db)}

@router.get("/collaborations-count")
def collaborations_count(db: Session = Depends(get_db)):
    return {"collaborations_count": crud.count_collaborations(db)}


@router.get("/generated")
def get_generated_reports(user: models.User = Depends(current_user), db: Session = Depends(get_db)):
    records = db.query(models.GeneratedReport)
    allowed_id = _allowed_institution_id(user, db)
    if allowed_id is not None:
        records = records.filter(models.GeneratedReport.institution_id == allowed_id)
    records = records.order_by(models.GeneratedReport.generated_at.desc()).all()
    return [
        {
            "institution_id": record.institution_id,
            "institution_name": record.institution.name,
            "researchers": record.researchers,
            "publications": record.publications,
            "generated_at": record.generated_at.isoformat() if record.generated_at else None,
        }
        for record in records
    ]


@router.post("/generated/{institution_id}")
def generate_institution_report(institution_id: int, user: models.User = Depends(require_roles("admin", "system admin", "institution admin")), db: Session = Depends(get_db)):
    _require_report_scope(institution_id, user, db)
    report = crud.get_institution_report(db, institution_id)
    if not report:
        raise HTTPException(status_code=404, detail="Institution not found")
    record = db.query(models.GeneratedReport).filter(models.GeneratedReport.institution_id == institution_id).first()
    if not record:
        record = models.GeneratedReport(institution_id=institution_id)
        db.add(record)
    record.researchers = report["researchers"]
    record.publications = report["publications"]
    record.generated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(record)
    notify_all_users(db, notification_type="report", title="Institution report generated", message=f"The analytics report for {report['institution_name']} is ready to view or export.", link=f"pages/institution-report.html?id={institution_id}")
    return {**report, "generated_at": record.generated_at.isoformat()}


@router.delete("/generated")
def clear_generated_reports(_admin: models.User = Depends(require_system_admin), db: Session = Depends(get_db)):
    db.query(models.GeneratedReport).delete()
    db.commit()
    return {"message": "Generated reports cleared"}


REPORT_STATUSES = ("published", "submitted", "draft", "archived")


def publication_status_by_year(publications):
    yearly = defaultdict(lambda: {status: 0 for status in REPORT_STATUSES})
    for publication in publications:
        if not publication.publication_date:
            continue
        status = (publication.status or "draft").lower()
        if status in REPORT_STATUSES:
            yearly[str(publication.publication_date.year)][status] += 1
    return [
        {"year": year, "total": sum(counts.values()), **counts}
        for year, counts in sorted(yearly.items(), key=lambda item: int(item[0]))
    ]


def researcher_rows(researchers):
    return [
        {
            "id": researcher.id,
            "name": researcher.full_name,
            "institution_id": researcher.institution_id,
            "publications": len(researcher.publications),
        }
        for researcher in sorted(researchers, key=lambda item: (-len(item.publications), item.full_name.lower()))
    ]


@router.get("/analytics")
def analytics_overview(institution_id: Optional[int] = Query(default=None), user: models.User = Depends(current_user), db: Session = Depends(get_db)):
    allowed_id = _allowed_institution_id(user, db)
    if allowed_id is not None:
        institution_id = allowed_id
    institution_options = db.query(models.Institution).order_by(models.Institution.name).all()
    if institution_id is not None and not any(institution.id == institution_id for institution in institution_options):
        raise HTTPException(status_code=404, detail="Institution not found")

    institutions = [institution for institution in institution_options if institution_id is None or institution.id == institution_id]
    researchers_query = db.query(models.Researcher)
    publications_query = db.query(models.Publication)
    if institution_id is not None:
        researchers_query = researchers_query.filter(models.Researcher.institution_id == institution_id)
        publications_query = publications_query.filter(models.Publication.institution_id == institution_id)
    researchers = researchers_query.all()
    publications = publications_query.all()

    institution_rows = [
        {
            "id": institution.id,
            "name": institution.name,
            "researchers": sum(researcher.institution_id == institution.id for researcher in researchers),
            "publications": sum(publication.institution_id == institution.id for publication in publications),
        }
        for institution in institutions
    ]
    researcher_ids = {researcher.id for researcher in researchers}
    collaborations_query = db.query(models.Collaboration)
    if institution_id is not None:
        collaborations_query = collaborations_query.filter(
            (models.Collaboration.researcher1_id.in_(researcher_ids)) |
            (models.Collaboration.researcher2_id.in_(researcher_ids))
        )
    collaborations = collaborations_query.all() if institution_id is None or researcher_ids else []
    collaboration_counts = Counter()
    for collaboration in collaborations:
        collaboration_counts[collaboration.researcher1_id] += 1
        collaboration_counts[collaboration.researcher2_id] += 1
    collaboration_rows = [
        {"id": researcher.id, "name": researcher.full_name, "collaborations": collaboration_counts[researcher.id]}
        for researcher in sorted(researchers, key=lambda item: (-collaboration_counts[item.id], item.full_name.lower()))
    ]
    publication_types = Counter((publication.publication_type or "Other").strip() or "Other" for publication in publications)

    return {
        "scope": {"institution_id": institution_id, "institution_name": institutions[0].name if institution_id is not None else None},
        "institution_options": [{"id": institution.id, "name": institution.name} for institution in institution_options],
        "institutions": institution_rows,
        "publication_status_by_year": publication_status_by_year(publications),
        "publication_types": [{"type": name, "count": count} for name, count in publication_types.most_common()],
        "top_researchers": researcher_rows(researchers)[:10],
        "top_collaborators": collaboration_rows[:10],
        "totals": {
            "institutions": len(institutions),
            "researchers": len(researchers),
            "publications": len(publications),
            "collaborations": len(collaborations),
        },
    }


@router.get("/institutions/{institution_id}")
def institution_full_report(institution_id: int, user: models.User = Depends(current_user), db: Session = Depends(get_db)):
    _require_report_scope(institution_id, user, db)
    institution = db.query(models.Institution).filter(models.Institution.id == institution_id).first()
    if not institution:
        raise HTTPException(status_code=404, detail="Institution not found")

    researchers = db.query(models.Researcher).filter(models.Researcher.institution_id == institution_id).all()
    publications = db.query(models.Publication).filter(models.Publication.institution_id == institution_id).all()
    researcher_ids = {researcher.id for researcher in researchers}
    collaborations = db.query(models.Collaboration).filter(
        (models.Collaboration.researcher1_id.in_(researcher_ids)) |
        (models.Collaboration.researcher2_id.in_(researcher_ids))
    ).all() if researcher_ids else []
    collaboration_counts = Counter()
    for collaboration in collaborations:
        if collaboration.researcher1_id in researcher_ids:
            collaboration_counts[collaboration.researcher1_id] += 1
        if collaboration.researcher2_id in researcher_ids:
            collaboration_counts[collaboration.researcher2_id] += 1

    return {
        "institution": {"id": institution.id, "name": institution.name},
        "totals": {"researchers": len(researchers), "publications": len(publications), "collaborations": len(collaborations)},
        "publication_status_by_year": publication_status_by_year(publications),
        "top_researchers": researcher_rows(researchers)[:10],
        "top_collaborators": [
            {"id": researcher.id, "name": researcher.full_name, "collaborations": collaboration_counts[researcher.id]}
            for researcher in sorted(researchers, key=lambda item: (-collaboration_counts[item.id], item.full_name.lower()))[:10]
        ],
    }


def _export_rows(report: dict):
    rows = [
        ["Institution report", report["institution"]["name"]],
        ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        [], ["Metric", "Value"],
    ]
    rows.extend([[key.replace("_", " ").title(), value] for key, value in report["totals"].items()])
    rows.extend([[], ["Year", "Published", "Submitted", "Draft", "Archived", "Total"]])
    rows.extend([[row["year"], row["published"], row["submitted"], row["draft"], row["archived"], row["total"]] for row in report["publication_status_by_year"]])
    rows.extend([[], ["Top researchers", "Publications"]])
    rows.extend([[row["name"], row["publications"]] for row in report["top_researchers"]])
    return rows


@router.get("/institutions/{institution_id}/export.xlsx")
def export_institution_excel(institution_id: int, user: models.User = Depends(current_user), db: Session = Depends(get_db)):
    """Download an actual Excel workbook, not a CSV file renamed as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    report = institution_full_report(institution_id, user, db)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Institution report"
    for row in _export_rows(report):
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E4A86")
    sheet.column_dimensions["A"].width = 28
    for column in "BCDEF":
        sheet.column_dimensions[column].width = 16
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"institution-report-{institution_id}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/institutions/{institution_id}/export.pdf")
def export_institution_pdf(institution_id: int, user: models.User = Depends(current_user), db: Session = Depends(get_db)):
    """Create a server-side PDF snapshot suitable for sharing or submission."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

    report = institution_full_report(institution_id, user, db)
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = [Paragraph("Institution Analytics Report", styles["Title"]), Paragraph(report["institution"]["name"], styles["Heading2"]), Spacer(1, 12)]
    totals = [["Metric", "Value"], *[[key.replace("_", " ").title(), str(value)] for key, value in report["totals"].items()]]
    yearly = [["Year", "Published", "Submitted", "Draft", "Archived", "Total"], *[[row["year"], row["published"], row["submitted"], row["draft"], row["archived"], row["total"]] for row in report["publication_status_by_year"]]]
    for heading, data in [("Summary", totals), ("Publication status by year", yearly)]:
        story.extend([Paragraph(heading, styles["Heading2"]), Table(data, repeatRows=1, hAlign="LEFT", style=TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E4A86")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D7DFEA")), ("PADDING", (0, 0), (-1, -1), 6)])), Spacer(1, 14)])
    document.build(story)
    output.seek(0)
    return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="institution-report-{institution_id}.pdf"'})
