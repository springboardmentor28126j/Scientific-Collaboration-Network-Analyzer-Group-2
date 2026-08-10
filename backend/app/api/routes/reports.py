import io
from typing import Literal

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.collaboration import Collaboration, CollaborationRequest, CollaborationRequestStatus
from app.models.conference import Conference
from app.models.institution import Institution
from app.models.project import Project, ProjectMember, ProjectStatus
from app.models.publication import Publication, PublicationAuthor, PublicationStatus
from app.models.researcher import Researcher
from app.models.user import User, UserRole
from app.schemas.report import (
    CollaborationReportOut,
    CountByKey,
    InstitutionReportOut,
    InstitutionReportRow,
    ProjectReportOut,
    PublicationReportOut,
    SummaryReportOut,
    TopCollaborationOut,
)

router = APIRouter()


def _current_institution_id_for_admin(db: Session, current_user: User) -> int | None:
    """The single institution an Institution Admin manages, or None if
    they don't manage one yet."""
    institution = (
        db.query(Institution).filter(Institution.admin_user_id == current_user.id).first()
    )
    return institution.id if institution else None


def _scope(db: Session, current_user: User) -> tuple[str, int | None, int | None]:
    """Returns (scope_label, institution_id_filter, researcher_id_filter).
    System Admin: no filters (sees everything).
    Institution Admin: filtered to their institution.
    Researcher / Reviewer: filtered to their own researcher_id.
    """
    if current_user.role == UserRole.SYSTEM_ADMIN:
        return "system", None, None
    if current_user.role == UserRole.INSTITUTION_ADMIN:
        institution_id = _current_institution_id_for_admin(db, current_user)
        return "institution", institution_id, None
    researcher = db.query(Researcher).filter(Researcher.user_id == current_user.id).first()
    return "researcher", None, (researcher.id if researcher else None)


@router.get("/summary", response_model=SummaryReportOut)
def summary_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> SummaryReportOut:
    scope, institution_id, researcher_id = _scope(db, current_user)

    pub_query = db.query(Publication)
    project_query = db.query(Project)

    if scope == "institution":
        if institution_id is None:
            pub_query = pub_query.filter(False)
            project_query = project_query.filter(False)
        else:
            pub_query = pub_query.join(PublicationAuthor).join(
                Researcher, Researcher.id == PublicationAuthor.researcher_id
            ).filter(Researcher.institution_id == institution_id)
            project_query = project_query.filter(Project.institution_id == institution_id)
    elif scope == "researcher":
        if researcher_id is None:
            pub_query = pub_query.filter(False)
            project_query = project_query.filter(False)
        else:
            pub_query = pub_query.join(PublicationAuthor).filter(
                PublicationAuthor.researcher_id == researcher_id
            )
            project_query = project_query.join(ProjectMember).filter(
                ProjectMember.researcher_id == researcher_id
            )

    publication_count = pub_query.distinct(Publication.id).count()
    published_count = pub_query.filter(Publication.status == PublicationStatus.PUBLISHED).distinct(
        Publication.id
    ).count()
    project_count = project_query.distinct(Project.id).count()
    active_project_count = project_query.filter(Project.status == ProjectStatus.ONGOING).distinct(
        Project.id
    ).count()

    out = SummaryReportOut(
        scope=scope,
        publication_count=publication_count,
        published_publication_count=published_count,
        project_count=project_count,
        active_project_count=active_project_count,
    )

    if scope == "system":
        out.researcher_count = db.query(Researcher).count()
        out.institution_count = db.query(Institution).count()
        out.conference_count = db.query(Conference).count()
        out.collaboration_count = db.query(Collaboration).count()
    elif scope == "institution" and institution_id is not None:
        out.researcher_count = (
            db.query(Researcher).filter(Researcher.institution_id == institution_id).count()
        )
        out.conference_count = (
            db.query(Conference).filter(Conference.institution_id == institution_id).count()
        )
    elif scope == "researcher" and researcher_id is not None:
        out.collaboration_count = (
            db.query(Collaboration)
            .filter(
                (Collaboration.researcher1_id == researcher_id)
                | (Collaboration.researcher2_id == researcher_id)
            )
            .count()
        )

    return out


@router.get("/publications", response_model=PublicationReportOut)
def publications_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PublicationReportOut:
    scope, institution_id, researcher_id = _scope(db, current_user)

    base = db.query(Publication)
    if scope == "institution":
        if institution_id is None:
            base = base.filter(False)
        else:
            base = base.join(PublicationAuthor).join(
                Researcher, Researcher.id == PublicationAuthor.researcher_id
            ).filter(Researcher.institution_id == institution_id)
    elif scope == "researcher":
        if researcher_id is None:
            base = base.filter(False)
        else:
            base = base.join(PublicationAuthor).filter(
                PublicationAuthor.researcher_id == researcher_id
            )

    ids_subquery = base.with_entities(Publication.id).distinct().subquery()
    scoped = db.query(Publication).filter(Publication.id.in_(ids_subquery))

    total = scoped.count()

    by_status = [
        CountByKey(key=row[0].value, count=row[1])
        for row in scoped.with_entities(Publication.status, func.count(Publication.id))
        .group_by(Publication.status)
        .all()
    ]
    by_type = [
        CountByKey(key=(row[0].value if row[0] else "unspecified"), count=row[1])
        for row in scoped.with_entities(Publication.type, func.count(Publication.id))
        .group_by(Publication.type)
        .all()
    ]
    by_year = [
        CountByKey(key=(str(row[0]) if row[0] else "unspecified"), count=row[1])
        for row in scoped.with_entities(Publication.year, func.count(Publication.id))
        .group_by(Publication.year)
        .order_by(Publication.year.desc().nullslast())
        .all()
    ]

    return PublicationReportOut(
        scope=scope, total=total, by_status=by_status, by_type=by_type, by_year=by_year
    )


@router.get("/projects", response_model=ProjectReportOut)
def projects_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ProjectReportOut:
    scope, institution_id, researcher_id = _scope(db, current_user)

    base = db.query(Project)
    if scope == "institution":
        base = base.filter(Project.institution_id == institution_id) if institution_id else base.filter(False)
    elif scope == "researcher":
        if researcher_id is None:
            base = base.filter(False)
        else:
            base = base.join(ProjectMember).filter(ProjectMember.researcher_id == researcher_id)

    ids_subquery = base.with_entities(Project.id).distinct().subquery()
    scoped = db.query(Project).filter(Project.id.in_(ids_subquery))

    total = scoped.count()
    by_status = [
        CountByKey(key=row[0].value, count=row[1])
        for row in scoped.with_entities(Project.status, func.count(Project.id))
        .group_by(Project.status)
        .all()
    ]
    return ProjectReportOut(scope=scope, total=total, by_status=by_status)


@router.get("/collaborations", response_model=CollaborationReportOut)
def collaborations_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> CollaborationReportOut:
    scope, institution_id, researcher_id = _scope(db, current_user)

    base = db.query(Collaboration)
    pending_base = db.query(CollaborationRequest).filter(
        CollaborationRequest.status == CollaborationRequestStatus.PENDING
    )
    if scope == "researcher" and researcher_id is not None:
        base = base.filter(
            (Collaboration.researcher1_id == researcher_id)
            | (Collaboration.researcher2_id == researcher_id)
        )
        pending_base = pending_base.filter(
            (CollaborationRequest.requester_id == researcher_id)
            | (CollaborationRequest.addressee_id == researcher_id)
        )
    elif scope == "institution" and institution_id is not None:
        r1 = db.query(Researcher.id).filter(Researcher.institution_id == institution_id).subquery()
        base = base.filter(
            Collaboration.researcher1_id.in_(r1) | Collaboration.researcher2_id.in_(r1)
        )

    total = base.count()
    avg_strength = base.with_entities(func.avg(Collaboration.strength)).scalar() or 0.0
    total_pending = pending_base.count()

    top_rows = (
        base.order_by(Collaboration.strength.desc()).limit(5).all()
    )
    top = [
        TopCollaborationOut(
            collaboration_id=c.id,
            researcher1_email=c.researcher1.user.email if c.researcher1.user else "",
            researcher2_email=c.researcher2.user.email if c.researcher2.user else "",
            strength=c.strength,
        )
        for c in top_rows
    ]

    return CollaborationReportOut(
        scope=scope,
        total_collaborations=total,
        total_pending_requests=total_pending,
        average_strength=round(float(avg_strength), 2),
        top_collaborations=top,
    )


@router.get("/institutions", response_model=InstitutionReportOut)
def institutions_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> InstitutionReportOut:
    if current_user.role not in (UserRole.SYSTEM_ADMIN, UserRole.INSTITUTION_ADMIN):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only a System Admin or Institution Admin can view institution reports",
        )

    query = db.query(Institution)
    if current_user.role == UserRole.INSTITUTION_ADMIN:
        institution_id = _current_institution_id_for_admin(db, current_user)
        query = query.filter(Institution.id == institution_id) if institution_id else query.filter(False)

    rows = []
    for institution in query.all():
        researcher_count = (
            db.query(Researcher).filter(Researcher.institution_id == institution.id).count()
        )
        publication_count = (
            db.query(Publication.id)
            .join(PublicationAuthor)
            .join(Researcher, Researcher.id == PublicationAuthor.researcher_id)
            .filter(Researcher.institution_id == institution.id)
            .distinct()
            .count()
        )
        project_count = (
            db.query(Project).filter(Project.institution_id == institution.id).count()
        )
        rows.append(
            InstitutionReportRow(
                institution_id=institution.id,
                name=institution.name,
                researcher_count=researcher_count,
                publication_count=publication_count,
                project_count=project_count,
            )
        )

    return InstitutionReportOut(rows=rows)


# -----------------------------
# Export (Module 8: Reports & Export -- Excel / PDF)
# -----------------------------

ReportType = Literal["summary", "publications", "projects", "collaborations", "institutions"]


def _flatten_report(report_type: str, current_user: User, db: Session) -> tuple[str, list[str], list[list]]:
    """Returns (title, headers, rows) for the given report type, ready to
    hand to either the Excel or PDF builder below. Reuses the same
    aggregation functions the JSON endpoints use, so exports always match
    what's on screen."""
    if report_type == "summary":
        data = summary_report(current_user, db)
        headers = ["Metric", "Value"]
        rows = [
            ["Scope", data.scope],
            ["Researchers", data.researcher_count],
            ["Institutions", data.institution_count],
            ["Publications", data.publication_count],
            ["Published Publications", data.published_publication_count],
            ["Projects", data.project_count],
            ["Active Projects", data.active_project_count],
            ["Conferences", data.conference_count],
            ["Collaborations", data.collaboration_count],
        ]
        return "Summary Report", headers, rows

    if report_type == "publications":
        data = publications_report(current_user, db)
        headers = ["Breakdown", "Key", "Count"]
        rows = [["Total", "", data.total]]
        rows += [["By Status", c.key, c.count] for c in data.by_status]
        rows += [["By Type", c.key, c.count] for c in data.by_type]
        rows += [["By Year", c.key, c.count] for c in data.by_year]
        return "Publication Report", headers, rows

    if report_type == "projects":
        data = projects_report(current_user, db)
        headers = ["Breakdown", "Key", "Count"]
        rows = [["Total", "", data.total]]
        rows += [["By Status", c.key, c.count] for c in data.by_status]
        return "Project Report", headers, rows

    if report_type == "collaborations":
        data = collaborations_report(current_user, db)
        headers = ["Metric", "Value"]
        rows = [
            ["Total Collaborations", data.total_collaborations],
            ["Pending Requests", data.total_pending_requests],
            ["Average Strength", data.average_strength],
        ]
        for t in data.top_collaborations:
            rows.append(
                [f"Top: {t.researcher1_email} <> {t.researcher2_email}", t.strength]
            )
        return "Collaboration Report", headers, rows

    if report_type == "institutions":
        data = institutions_report(current_user, db)
        headers = ["Institution", "Researchers", "Publications", "Projects"]
        rows = [
            [r.name, r.researcher_count, r.publication_count, r.project_count]
            for r in data.rows
        ]
        return "Institution Report", headers, rows

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown report type")


def _build_excel(title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)

    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_pdf(title: str, headers: list[str], rows: list[list]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    table_data = [headers] + [[str(cell) for cell in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d3748")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/export/excel")
def export_excel(
    type: ReportType,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    title, headers, rows = _flatten_report(type, current_user, db)
    content = _build_excel(title, headers, rows)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{type}_report.xlsx"'},
    )


@router.get("/export/pdf")
def export_pdf(
    type: ReportType,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    title, headers, rows = _flatten_report(type, current_user, db)
    content = _build_pdf(title, headers, rows)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{type}_report.pdf"'},
    )
