import tempfile
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def _report_output_path(filename: str) -> str:
    """Writes generated report files to the OS temp directory instead of the
    backend's working directory, using a random suffix so two people
    downloading a report at the same time never read/write the same file."""
    unique_name = f"{filename.rsplit('.', 1)[0]}_{uuid.uuid4().hex[:8]}.{filename.rsplit('.', 1)[1]}"
    return str(Path(tempfile.gettempdir()) / unique_name)


def create_dashboard_excel(summary, collaboration_status=None, top_collaborations=None,
                            institution_report=None, publication_year=None, publication_type=None,
                            publication_status=None, conference_type=None, user_roles=None,
                            departments=None, interests=None, skills=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard Report"
    sheet.append(["Metric", "Value"])
    bold = Font(bold=True)
    sheet["A1"].font = bold
    sheet["B1"].font = bold

    data = [
        ("Total Users", summary.total_users),
        ("Total Researchers", summary.total_researchers),
        ("Total Institutions", summary.total_institutions),
        ("Total Publications", summary.total_publications),
        ("Total Conferences", summary.total_conferences),
        ("Total Sessions", summary.total_sessions),
        ("Total Participations", summary.total_participations),
        ("Total Collaborations", summary.total_collaborations),
    ]
    for row in data:
        sheet.append(row)

    def _add_sheet(name, headers, rows, fields):
        if not rows:
            return
        ws = workbook.create_sheet(name)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = bold
        for row in rows:
            ws.append([getattr(row, f) for f in fields])

    if collaboration_status:
        _add_sheet("Collaboration Requests", ["Status", "Count"], collaboration_status,
                    ["request_status", "request_count"])
    if top_collaborations:
        _add_sheet("Top Collaborations", ["Researcher 1", "Researcher 2", "Shared Publications"],
                    top_collaborations, ["researcher1_email", "researcher2_email", "strength"])
    if institution_report:
        _add_sheet("Institutions", ["Institution", "Researchers"], institution_report,
                    ["institution_name", "researcher_count"])
    if publication_year:
        _add_sheet("Publications by Year", ["Year", "Count"], publication_year,
                    ["year", "publication_count"])
    if publication_type:
        _add_sheet("Publications by Type", ["Type", "Count"], publication_type,
                    ["publication_type", "publication_count"])
    if publication_status:
        _add_sheet("Publications by Status", ["Status", "Count"], publication_status,
                    ["publication_status", "publication_count"])
    if conference_type:
        _add_sheet("Conferences by Type", ["Type", "Count"], conference_type,
                    ["conference_type", "conference_count"])
    if user_roles:
        _add_sheet("Users by Role", ["Role", "Count"], user_roles, ["role", "total_users"])
    if departments:
        _add_sheet("Departments", ["Department", "Researchers"], departments,
                    ["department", "total_researchers"])
    if interests:
        _add_sheet("Research Interests", ["Interest", "Researchers"], interests,
                    ["research_interest", "total_researchers"])
    if skills:
        _add_sheet("Skills", ["Skill", "Researchers"], skills, ["skill", "total_researchers"])

    filename = _report_output_path("dashboard_report.xlsx")
    workbook.save(filename)
    return filename


def create_compliance_excel(report: dict) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compliance Summary"
    bold = Font(bold=True)

    sheet.append(["Compliance Report", f"{report['period']['start']} to {report['period']['end']}"])
    sheet["A1"].font = bold
    sheet.append([])
    sheet.append(["Metric", "Count"])
    sheet["A3"].font = bold
    sheet["B3"].font = bold
    for key, value in report["totals"].items():
        sheet.append([key.replace("_", " ").title(), value])

    decisions_ws = workbook.create_sheet("Publication Decisions")
    decisions_ws.append(["Date", "Action", "Reviewer", "Comment"])
    for col in range(1, 5):
        decisions_ws.cell(row=1, column=col).font = bold
    for row in report["publication_decisions"]:
        decisions_ws.append([row["date"], row["action"], row["actor_email"], row["details"]])

    failures_ws = workbook.create_sheet("Login Failures")
    failures_ws.append(["Date", "Attempted Email"])
    for col in range(1, 3):
        failures_ws.cell(row=1, column=col).font = bold
    for row in report["login_failures"]:
        failures_ws.append([row["date"], row["attempted_email"]])

    mfa_ws = workbook.create_sheet("MFA Events")
    mfa_ws.append(["Date", "Action", "User"])
    for col in range(1, 4):
        mfa_ws.cell(row=1, column=col).font = bold
    for row in report["mfa_events"]:
        mfa_ws.append([row["date"], row["action"], row["actor_email"]])

    filename = _report_output_path("compliance_report.xlsx")
    workbook.save(filename)
    return filename