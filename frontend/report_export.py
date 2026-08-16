"""
Shared export machinery for the /reports pages. Every report already
computes its own data for the HTML render; this module just turns that
same data into a downloadable .xlsx or .pdf. Each report builds a
ReportSection list (one section = one table) and a summary key/value
list, then calls build_excel() or build_pdf() -- the report routes never
touch openpyxl/reportlab directly.
"""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

BRAND_INK = "1C2541"
BRAND_ACCENT = "3457D5"
BRAND_BORDER = "E3E6EF"


class ReportSection:
    """One table within a report -- becomes one worksheet in the xlsx,
    one titled table in the pdf. rows should already be display-ready
    strings/numbers (format dates, booleans, etc. before building this)."""

    def __init__(self, title: str, headers: list[str], rows: list[list]):
        self.title = title
        self.headers = headers
        self.rows = rows


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def build_excel(report_title: str, summary: list[tuple[str, str]], sections: list[ReportSection]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet, we add our own named ones

    header_fill = PatternFill(start_color=BRAND_INK, end_color=BRAND_INK, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if summary:
        ws = wb.create_sheet("Summary")
        ws["A1"] = report_title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Generated {_timestamp()}"
        ws["A2"].font = Font(italic=True, size=9, color="666666")
        row = 4
        for label, value in summary:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 26

    for section in sections:
        # Excel sheet names: 31 char max, and these characters aren't allowed at all.
        safe_name = "".join(c for c in section.title if c not in '[]:*?/\\')[:31] or "Sheet"
        # openpyxl requires unique sheet names; disambiguate on collision.
        base_name, n = safe_name, 1
        while safe_name in wb.sheetnames:
            n += 1
            safe_name = f"{base_name[:28]} {n}"
        ws = wb.create_sheet(safe_name)

        for col_idx, header in enumerate(section.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        for row_idx, row_data in enumerate(section.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, header in enumerate(section.headers, start=1):
            candidates = [len(str(header))] + [len(str(r[col_idx - 1])) for r in section.rows if col_idx - 1 < len(r)]
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(candidates) + 2, 50)

        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("Report")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(report_title: str, summary: list[tuple[str, str]], sections: list[ReportSection]) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], textColor=colors.HexColor(f"#{BRAND_INK}"))
    section_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"], textColor=colors.HexColor(f"#{BRAND_ACCENT}"), spaceBefore=16
    )
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], textColor=colors.HexColor("#666666"), fontSize=9)

    story = [Paragraph(report_title, title_style), Paragraph(f"Generated {_timestamp()}", meta_style), Spacer(1, 14)]

    if summary:
        data = [[label, str(value)] for label, value in summary]
        t = Table(data, colWidths=[2.3 * inch, 3.9 * inch])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BRAND_BORDER}")),
        ]))
        story.append(t)
        story.append(Spacer(1, 16))

    for section in sections:
        story.append(Paragraph(section.title, section_style))
        if not section.rows:
            story.append(Paragraph("No data.", styles["Normal"]))
            continue

        # Long tables need to wrap cell text, not just overflow -- wrap
        # every cell in a Paragraph so reportlab handles line-breaking
        # instead of the table just running off the page edge.
        cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)
        header_style = ParagraphStyle("CellHeader", parent=cell_style, textColor=colors.white, fontName="Helvetica-Bold")
        data = [[Paragraph(str(h), header_style) for h in section.headers]]
        for row in section.rows:
            data.append([Paragraph(str(v) if v is not None else "", cell_style) for v in row])

        col_count = len(section.headers)
        avail_width = 7.3 * inch
        col_width = avail_width / col_count
        t = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_INK}")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BRAND_BORDER}")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6FB")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    doc.build(story)
    buf.seek(0)
    return buf
